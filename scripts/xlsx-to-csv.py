import csv
import io
import sys
from openpyxl import load_workbook


def main():
    if len(sys.argv) < 2:
        print("Missing xlsx path", file=sys.stderr)
        return 1

    workbook = load_workbook(filename=sys.argv[1], data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")

    for row in sheet.iter_rows(values_only=True):
        values = []
        for cell in row:
            values.append("" if cell is None else str(cell))
        writer.writerow(values)

    sys.stdout.write(output.getvalue())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
